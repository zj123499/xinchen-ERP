#!/usr/bin/env python3
"""
新辰ERP ETL导入引擎 v1.0
=====================
配置驱动的Excel→PostgreSQL数据迁移流水线

用法:
  python3 engine.py --config mapping.json --dry-run     # 预览模式
  python3 engine.py --config mapping.json                # 正式导入
  python3 engine.py --config mapping.json --source lead  # 只导指定表
  python3 engine.py --config mapping.json --report       # 只生成报告

配置文件格式 mapping.json:

{
  "global": {
    "tenantId": 1,
    "dataDir": "/path/to/excel/files",
    "batchSize": 200
  },
  "sources": {
    "sourceName": {
      "label": "显示名称",
      "files": ["文件名_*.xlsx"],
      "target": "目标数据库表名",
      "priority": 1,
      "fields": [
        {"source": "Excel列名", "target": "DB字段名",
         "type": "string|int|float|date|datetime|phone",
         "required": true,
         "default": "默认值",
         "enum": {"原值": "目标值"},
         "lookup": {"table": "表名", "key": "匹配字段", "return": "id"},
         "concat": ["其它列名"],      // 拼接多个列的值
         "ignore": "skip if empty"    // 空值跳过
        }
      ]
    }
  }
}

支持的 field.type:
  string   - 字符串
  phone    - 清理 +86/86前缀, 截断至20字符
  int      - 整数
  float    - 浮点
  date     - 日期 (YYYY-MM-DD)
  datetime - 日期时间 (YYYY-MM-DD HH:MM:SS)
  bool     - 布尔

支持的 field.transform:
  enum     - 枚举值映射 (需配 field.map)
  lookup   - 外键查找 (需配 field.lookup)
  concat   - 多列拼接 (需配 field.concat)
"""

import os, sys, json, re, glob, time
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("[ERROR] 需要 openpyxl: pip3 install openpyxl")
    sys.exit(1)
try:
    import psycopg2
except ImportError:
    print("[ERROR] 需要 psycopg2: pip3 install psycopg2-binary")
    sys.exit(1)


# ====================================================================
# 数据提取层 (Extract)
# ====================================================================

class ExcelExtractor:
    """从 Excel 文件提取原始数据"""

    def __init__(self, data_dir):
        self.data_dir = data_dir

    def read(self, file_pattern):
        """读取匹配模式的所有 Excel，使用第2行(中文)作为列名"""
        full_pattern = os.path.join(self.data_dir, file_pattern)
        matches = glob.glob(full_pattern)
        if not matches:
            print(f"  [WARN] 未找到文件: {full_pattern}")
            return []

        all_rows = []
        for fp in sorted(matches):
            wb = openpyxl.load_workbook(fp, data_only=True)
            ws = wb["Sheet1"]
            # 第2行=中文列名, 第3行起=数据
            headers = [str(c.value or "").strip() for c in ws[2]]
            for row in ws.iter_rows(min_row=3, values_only=True):
                record = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        record[headers[i]] = str(val) if val is not None else ""
                if record:
                    all_rows.append(record)
        return all_rows


# ====================================================================
# 数据转换层 (Transform)
# ====================================================================

class FieldTransformer:
    """字段级转换器 — 支持类型转换、枚举映射、外键查找、多列拼接"""

    def __init__(self, conn, tenant_id):
        self.conn = conn
        self.tenant_id = tenant_id
        self._lookup_cache = {}  # (table, key, value) → id

    def transform(self, raw_value, field_config):
        """对单个字段值执行完整转换"""
        val = raw_value.strip() if isinstance(raw_value, str) else str(raw_value) if raw_value is not None else ""
        ftype = field_config.get("type", "string")
        is_default = False

        # 空值处理
        if not val:
            if field_config.get("default") is not None:
                val = field_config["default"]
                is_default = True
            if field_config.get("required"):
                return None
            if not is_default:
                return None

        # 默认值也需要类型转换
        if is_default and ftype == "datetime":
            if val == "now":
                return datetime.now()
            try:
                return datetime.fromisoformat(str(val))
            except:
                return datetime.now()
        if is_default and ftype == "int":
            return int(val) if val is not None else 0
        if is_default and ftype == "float":
            return float(val) if val is not None else 0.0
        if is_default and ftype == "bool":
            return bool(val)
        if is_default and ftype == "string":
            return str(val) if val is not None else ""

        # 类型转换
        if ftype == "phone":
            val = re.sub(r"^\+86|^86", "", val)
            val = re.sub(r"[^\d]", "", val)[:20]
        elif ftype in ("int", "integer"):
            try:
                val = int(float(val.replace(",", "")))
            except:
                return None
        elif ftype == "float":
            try:
                val = float(val.replace(",", ""))
            except:
                return None
        elif ftype == "date":
            try:
                val = datetime.strptime(val[:10], "%Y-%m-%d")
            except:
                try:
                    val = datetime.strptime(val[:10], "%Y/%m/%d")
                except:
                    return None
        elif ftype == "datetime":
            try:
                val = datetime.strptime(val[:19], "%Y-%m-%d %H:%M:%S")
            except:
                try:
                    val = datetime.strptime(val[:10], "%Y-%m-%d")
                except:
                    return None
        elif ftype == "bool":
            val = val in ("是", "true", "True", "1", "YES", "yes", "有")

        # 枚举值映射
        if "enum" in field_config and val in field_config["enum"]:
            val = field_config["enum"][val]

        # 外键查找
        if "lookup" in field_config:
            lk = field_config["lookup"]
            table = lk["table"]
            key = lk["key"]
            cache_k = (table, key, val)
            if cache_k in self._lookup_cache:
                val = self._lookup_cache[cache_k]
            else:
                cur = self.conn.cursor()
                cur.execute(
                    f"SELECT id FROM {table} WHERE {key}=%s AND tenant_id=%s LIMIT 1",
                    (val, self.tenant_id),
                )
                row = cur.fetchone()
                val = row[0] if row else None
                self._lookup_cache[cache_k] = val

        return val

    def concat_fields(self, record, field_names, sep=" | "):
        """拼接多个字段的值"""
        parts = []
        for fn in field_names:
            v = record.get(fn, "").strip()
            if v:
                parts.append(v)
        return sep.join(parts) if parts else None


class RecordTransformer:
    """记录级转换器 — 将原始 Excel 行转换为目标 DB 行"""

    def __init__(self, conn, tenant_id):
        self.conn = conn
        self.tenant_id = tenant_id
        self.fields = FieldTransformer(conn, tenant_id)

    def transform_one(self, raw, field_configs):
        """转换单条记录"""
        result = {"tenant_id": self.tenant_id}

        for fc in field_configs:
            target = fc["target"]

            # 多列拼接
            if fc.get("type") == "concat" or "concat" in fc:
                concat_cols = fc.get("concat", fc.get("columns", []))
                val = self.fields.concat_fields(raw, concat_cols)
                if val:
                    result[target] = val
                continue

            # 单列映射
            source = fc.get("source", fc.get("name", ""))
            raw_val = raw.get(source, "")

            # 跳过静默列
            if fc.get("ignore") and not raw_val.strip():
                continue

            val = self.fields.transform(raw_val, fc)

            if val is not None:
                result[target] = val
            elif fc.get("required"):
                return None, f"必填字段 {source} 为空"

        # 自动添加时间戳
        if "created_at" not in result:
            result["created_at"] = datetime.now()
        if "updated_at" not in result:
            result["updated_at"] = datetime.now()

        return result, None


# ====================================================================
# 数据加载层 (Load)
# ====================================================================

class BatchLoader:
    """批量写入 PostgreSQL"""

    def __init__(self, conn, batch_size=200):
        self.conn = conn
        self.batch_size = batch_size
        self.stats = {"inserted": 0, "skipped": 0, "errors": 0}
        self.errors = []

    def upsert(self, table, records, unique_keys=None, dry_run=False):
        """批量 upsert 写入
        - unique_keys: 用于 ON CONFLICT 的唯一键列表
        - dry_run: 仅统计不写入
        """
        if not records:
            return

        total = len(records)
        self.stats["inserted"] += total

        if dry_run:
            return

        columns = list(records[0].keys())
        placeholders = ", ".join(["%s"] * len(columns))
        col_str = ", ".join(columns)

        conflict_clause = "DO NOTHING"
        if unique_keys:
            uk_str = ", ".join(unique_keys)
            conflict_clause = f"ON CONFLICT ({uk_str}) DO NOTHING"

        sql = f'INSERT INTO {table} ({col_str}) VALUES ({placeholders}) {conflict_clause}'

        cur = self.conn.cursor()
        try:
            for i in range(0, total, self.batch_size):
                batch = records[i : i + self.batch_size]
                values = [tuple(r[c] for c in columns) for r in batch]
                cur.executemany(sql, values)
                self.conn.commit()
        except Exception as e:
            self.conn.rollback()
            self.stats["errors"] += total
            self.stats["inserted"] -= total
            self.errors.append(f"  [{table}] 批量写入失败: {e}")
        finally:
            cur.close()


# ====================================================================
# 数据校验层 (Validate)
# ====================================================================

class DataValidator:
    """导入前后数据校验"""

    def __init__(self, conn, tenant_id):
        self.conn = conn
        self.tenant_id = tenant_id

    def pre_check(self, rows, field_configs):
        """导入前检查：统计空值率、重复率"""
        report = {"total": len(rows), "warnings": []}

        for fc in field_configs:
            source = fc.get("source", fc.get("name", ""))
            if not source:
                continue
            empty = sum(1 for r in rows if not r.get(source, "").strip())
            target = fc.get("target", "?")
            pct = round(empty / max(len(rows), 1) * 100, 1)
            if pct > 30 and not fc.get("default"):
                report["warnings"].append(f"  [{target}] 空值率 {pct}% ({empty}/{len(rows)})")

        return report

    def post_check(self, table, before_count):
        """导入后检查：数量是否正确"""
        cur = self.conn.cursor()
        cur.execute(f"SELECT count(*) FROM {table} WHERE tenant_id=%s", (self.tenant_id,))
        after = cur.fetchone()[0]
        return after - before_count


# ====================================================================
# 报告生成层 (Report)
# ====================================================================

class ReportGenerator:
    """生成导入报告"""

    def __init__(self, results):
        self.results = results  # {source_name: {stats, warnings, errors}}

    def print_summary(self):
        print("\n" + "=" * 60)
        print("  ETL 导入报告")
        print("=" * 60)

        total_in, total_skip, total_err = 0, 0, 0
        for name, r in sorted(self.results.items()):
            s = r.get("stats", {})
            ri = s.get("inserted", 0)
            rs = s.get("skipped", 0)
            re = s.get("errors", 0)
            total_in += ri
            total_skip += rs
            total_err += re
            status = "PASS" if re == 0 else "FAIL"
            print(f"  [{status}] {r.get('label', name):20s}  "
                  f"导入:{ri:>5}  跳过:{rs:>4}  错误:{re:>4}")

        print(f"  {'─' * 52}")
        print(f"  TOTAL                             "
              f"导入:{total_in:>5}  跳过:{total_skip:>4}  错误:{total_err:>4}")

        # 警告
        all_warnings = []
        for name, r in self.results.items():
            for w in r.get("warnings", []):
                all_warnings.append(f"  [{name}] {w}")
        if all_warnings:
            print(f"\n  数据质量警告:")
            for w in all_warnings[:10]:
                print(w)

        # 错误
        all_errors = []
        for name, r in self.results.items():
            for e in r.get("errors", []):
                all_errors.append(e)
        if all_errors:
            print(f"\n  错误详情:")
            for e in all_errors:
                print(e)

        print("=" * 60)

    def to_markdown(self, filepath):
        """导出 Markdown 报告"""
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"# ETL 导入报告\n\n")
            f.write(f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(f"| 数据源 | 状态 | 导入 | 跳过 | 错误 |\n")
            f.write(f"|--------|------|------|------|------|\n")
            for name, r in sorted(self.results.items()):
                s = r.get("stats", {})
                status = "PASS" if s.get("errors", 0) == 0 else "FAIL"
                f.write(f"| {r.get('label', name)} | {status} | "
                        f"{s.get('inserted', 0)} | {s.get('skipped', 0)} | "
                        f"{s.get('errors', 0)} |\n")
            total_in = sum(r.get("stats", {}).get("inserted", 0) for r in self.results.values())
            f.write(f"\n**总计导入: {total_in} 条**\n")


# ====================================================================
# 流水线编排 (Pipeline)
# ====================================================================

class ETLPipeline:
    """ETL 流水线主控制器"""

    def __init__(self, config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            self.config = json.load(f)
        self.global_cfg = self.config.get("global", {})
        self.tenant_id = self.global_cfg.get("tenantId", 1)
        self.data_dir = self.global_cfg.get("dataDir", ".")
        self.batch_size = self.global_cfg.get("batchSize", 200)
        self.dry_run = False
        self.source_filter = None
        self.results = {}

    def run(self, dry_run=False, source_filter=None, clean_first=False):
        self.dry_run = dry_run
        self.source_filter = source_filter

        # 数据库连接
        db_cfg = self.global_cfg.get("database", {})
        conn = psycopg2.connect(
            host=db_cfg.get("host", "127.0.0.1"),
            port=db_cfg.get("port", 5432),
            dbname=db_cfg.get("dbname", "xinchen_erp"),
            user=db_cfg.get("user", "user_TJ4skN"),
            password=db_cfg.get("password", "password_6Jaz2n"),
        )

        extractor = ExcelExtractor(self.data_dir)
        transformer = RecordTransformer(conn, self.tenant_id)
        loader = BatchLoader(conn, self.batch_size)
        validator = DataValidator(conn, self.tenant_id)

        sources = self.config.get("sources", {})
        sorted_sources = sorted(sources.items(), key=lambda x: x[1].get("priority", 99))

        for name, cfg in sorted_sources:
            if self.source_filter and name != self.source_filter:
                continue

            label = cfg.get("label", name)
            target_table = cfg.get("target", name)
            field_configs = cfg.get("fields", [])
            files = cfg.get("files", [])
            unique_keys = cfg.get("uniqueKeys", None)

            print(f"\n{'[DRY-RUN] ' if dry_run else ''}[{label}] → {target_table}")

            # 1. Extract
            all_rows = []
            for fp in files:
                rows = extractor.read(fp)
                print(f"  提取: {os.path.basename(fp)} → {len(rows)} 行")
                all_rows.extend(rows)
            print(f"  总计: {len(all_rows)} 行")

            # 2. Pre-validate
            pre = validator.pre_check(all_rows, field_configs)

            # 3. Clean (可选)
            if clean_first and not dry_run:
                cur = conn.cursor()
                clean_cond = cfg.get("cleanCondition", "source='IMPORT' AND tenant_id=%s")
                cur.execute(f"DELETE FROM {target_table} WHERE {clean_cond}", (self.tenant_id,))
                conn.commit()

            # 4. Transform
            transformed = []
            for row in all_rows:
                record, err = transformer.transform_one(row, field_configs)
                if err:
                    loader.errors.append(f"  [{label}] transform error: {err} from row={row.get(field_configs[0]['source'],'?')}")
                    loader.stats["skipped"] += 1
                elif record:
                    transformed.append(record)

            if not transformed:
                print(f"  转换: 0 条有效记录")
                self.results[name] = {
                    "label": label,
                    "stats": loader.stats.copy(),
                    "warnings": pre.get("warnings", []),
                    "errors": loader.errors,
                }
                continue

            # 5. Load
            before_count = 0
            if not dry_run:
                before_count = validator.post_check(target_table, 0)  # just check count
                cur = conn.cursor()
                cur.execute(f"SELECT count(*) FROM {target_table} WHERE tenant_id=%s", (self.tenant_id,))
                before_count = cur.fetchone()[0]

            print(f"  转换: {len(transformed)} 条 → ", end="")
            loader.upsert(target_table, transformed, unique_keys, dry_run)
            print(f"写入 {loader.stats['inserted'] - sum(self.results.get(name, {'stats': {'inserted': 0}})['stats'].get('inserted', 0) for _ in [0])} 条")

            # 6. Post-validate
            added = 0
            if not dry_run:
                added = validator.post_check(target_table, before_count)
                print(f"  校验: 实际新增 {added} 条")

            self.results[name] = {
                "label": label,
                "stats": loader.stats.copy(),
                "warnings": pre.get("warnings", []),
                "errors": loader.errors,
                "added": added,
            }

        conn.close()

        # 生成报告
        report = ReportGenerator(self.results)
        report.print_summary()

        md_path = self.global_cfg.get("reportPath", "etl_report.md")
        report.to_markdown(md_path)
        print(f"\n报告已保存: {md_path}")

        return self.results


# ====================================================================
# CLI 入口
# ====================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="新辰ERP ETL导入引擎")
    parser.add_argument("--config", required=True, help="映射配置文件路径 (JSON)")
    parser.add_argument("--dry-run", action="store_true", help="预览模式：不写入数据库")
    parser.add_argument("--source", help="只导入指定数据源 (如: lead)")
    parser.add_argument("--report-only", action="store_true", help="只生成报告")
    parser.add_argument("--clean", action="store_true", help="导入前清理现有脏数据")
    args = parser.parse_args()

    if args.report_only:
        # 只生成空报告模板
        print("报告模式 - 检查配置文件...")
        with open(args.config) as f:
            cfg = json.load(f)
        sources = cfg.get("sources", {})
        for name, sc in sorted(sources.items()):
            label = sc.get("label", name)
            files = sc.get("files", [])
            print(f"  [{label}] → {sc.get('target')} (文件: {', '.join(files)})")
        sys.exit(0)

    pipeline = ETLPipeline(args.config)
    pipeline.run(
        dry_run=args.dry_run,
        source_filter=args.source,
        clean_first=args.clean,
    )
